#!/usr/bin/env python3
"""Generate DXF test fixtures using openpyxl.

This script creates XLSX files with various DXF (Differential Format) configurations
for testing duke-sheets' DXF parsing and writing capabilities.

Usage:
    python3 generate_fixtures.py

Requirements:
    pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import os

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))


def create_dxf_font_basic_test():
    """Test: DXF with basic font properties (bold, italic, color)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "FontBasic"

    # Add test data
    for i in range(1, 6):
        ws.cell(row=i, column=1, value=i * 10)

    # CF rule with bold red font
    dxf = DifferentialStyle(font=Font(bold=True, color="FF0000"))
    rule = Rule(type="cellIs", operator="greaterThan", formula=["30"], dxf=dxf)
    ws.conditional_formatting.add("A1:A5", rule)

    wb.save(os.path.join(OUTPUT_DIR, "dxf_font_basic.xlsx"))
    print("Created: dxf_font_basic.xlsx")


def create_dxf_font_effects_test():
    """Test: DXF with various font effects."""
    wb = Workbook()
    ws = wb.active
    ws.title = "FontEffects"

    for i in range(1, 6):
        ws.cell(row=i, column=1, value=f"Value {i}")

    dxf = DifferentialStyle(
        font=Font(
            bold=True, italic=True, underline="single", strike=True, color="0000FF"
        )
    )
    rule = Rule(type="cellIs", operator="equal", formula=['"Value 3"'], dxf=dxf)
    ws.conditional_formatting.add("A1:A5", rule)

    wb.save(os.path.join(OUTPUT_DIR, "dxf_font_effects.xlsx"))
    print("Created: dxf_font_effects.xlsx")


def create_dxf_fill_solid_test():
    """Test: DXF with solid fill."""
    wb = Workbook()
    ws = wb.active
    ws.title = "FillSolid"

    for i in range(1, 6):
        ws.cell(row=i, column=1, value=i * 20)

    dxf = DifferentialStyle(fill=PatternFill(start_color="FFFF00", fill_type="solid"))
    rule = Rule(type="cellIs", operator="greaterThan", formula=["50"], dxf=dxf)
    ws.conditional_formatting.add("A1:A5", rule)

    wb.save(os.path.join(OUTPUT_DIR, "dxf_fill_solid.xlsx"))
    print("Created: dxf_fill_solid.xlsx")


def create_dxf_border_test():
    """Test: DXF with border styles."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Border"

    for i in range(1, 6):
        ws.cell(row=i, column=1, value=i * 100)

    dxf = DifferentialStyle(
        border=Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="medium", color="0000FF"),
            bottom=Side(style="medium", color="0000FF"),
        )
    )
    rule = Rule(type="cellIs", operator="greaterThan", formula=["200"], dxf=dxf)
    ws.conditional_formatting.add("A1:A5", rule)

    wb.save(os.path.join(OUTPUT_DIR, "dxf_border.xlsx"))
    print("Created: dxf_border.xlsx")


def create_dxf_alignment_test():
    """Test: DXF with alignment properties."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Alignment"

    for i in range(1, 6):
        ws.cell(row=i, column=1, value=f"Text {i}")

    dxf = DifferentialStyle(
        alignment=Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
            indent=2,
            text_rotation=45,
        )
    )
    rule = Rule(type="cellIs", operator="equal", formula=['"Text 3"'], dxf=dxf)
    ws.conditional_formatting.add("A1:A5", rule)

    wb.save(os.path.join(OUTPUT_DIR, "dxf_alignment.xlsx"))
    print("Created: dxf_alignment.xlsx")


def create_dxf_alignment_wrap_test():
    """Test: DXF with text wrapping."""
    wb = Workbook()
    ws = wb.active
    ws.title = "AlignWrap"

    for i in range(1, 6):
        ws.cell(row=i, column=1, value=f"This is a long text value {i}")

    dxf = DifferentialStyle(
        alignment=Alignment(horizontal="left", vertical="top", wrap_text=True)
    )
    rule = Rule(type="cellIs", operator="greaterThan", formula=["0"], dxf=dxf)
    ws.conditional_formatting.add("A1:A5", rule)

    wb.save(os.path.join(OUTPUT_DIR, "dxf_alignment_wrap.xlsx"))
    print("Created: dxf_alignment_wrap.xlsx")


def create_dxf_protection_test():
    """Test: DXF with protection properties."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Protection"

    for i in range(1, 6):
        ws.cell(row=i, column=1, value=i * 100)

    dxf = DifferentialStyle(protection=Protection(locked=False, hidden=True))
    rule = Rule(type="cellIs", operator="greaterThan", formula=["300"], dxf=dxf)
    ws.conditional_formatting.add("A1:A5", rule)

    wb.save(os.path.join(OUTPUT_DIR, "dxf_protection.xlsx"))
    print("Created: dxf_protection.xlsx")


def create_dxf_full_style_test():
    """Test: DXF with all style properties combined."""
    wb = Workbook()
    ws = wb.active
    ws.title = "FullStyle"

    for i in range(1, 11):
        ws.cell(row=i, column=1, value=i * 10)

    dxf = DifferentialStyle(
        font=Font(bold=True, italic=True, color="FF0000", size=14),
        fill=PatternFill(start_color="FFFF00", fill_type="solid"),
        border=Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="medium", color="0000FF"),
            bottom=Side(style="medium", color="0000FF"),
        ),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    rule = Rule(type="cellIs", operator="greaterThan", formula=["50"], dxf=dxf)
    ws.conditional_formatting.add("A1:A10", rule)

    wb.save(os.path.join(OUTPUT_DIR, "dxf_full_style.xlsx"))
    print("Created: dxf_full_style.xlsx")


def create_dxf_multiple_rules_test():
    """Test: Multiple CF rules with different DXF styles."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MultipleRules"

    for i in range(1, 11):
        ws.cell(row=i, column=1, value=i * 10)

    # Rule 1: Red fill for values > 70
    dxf1 = DifferentialStyle(
        fill=PatternFill(start_color="FF0000", fill_type="solid"),
        font=Font(color="FFFFFF", bold=True),
    )
    rule1 = Rule(
        type="cellIs", operator="greaterThan", formula=["70"], dxf=dxf1, priority=1
    )
    ws.conditional_formatting.add("A1:A10", rule1)

    # Rule 2: Yellow fill for values > 40
    dxf2 = DifferentialStyle(fill=PatternFill(start_color="FFFF00", fill_type="solid"))
    rule2 = Rule(
        type="cellIs", operator="greaterThan", formula=["40"], dxf=dxf2, priority=2
    )
    ws.conditional_formatting.add("A1:A10", rule2)

    # Rule 3: Green fill for values > 20
    dxf3 = DifferentialStyle(fill=PatternFill(start_color="00FF00", fill_type="solid"))
    rule3 = Rule(
        type="cellIs", operator="greaterThan", formula=["20"], dxf=dxf3, priority=3
    )
    ws.conditional_formatting.add("A1:A10", rule3)

    wb.save(os.path.join(OUTPUT_DIR, "dxf_multiple_rules.xlsx"))
    print("Created: dxf_multiple_rules.xlsx")


def create_dxf_formula_rule_test():
    """Test: DXF with formula-based rule."""
    wb = Workbook()
    ws = wb.active
    ws.title = "FormulaRule"

    # Create a grid of values
    for row in range(1, 6):
        for col in range(1, 4):
            ws.cell(row=row, column=col, value=row * col)

    # Rule: Highlight cells where value is even
    dxf = DifferentialStyle(fill=PatternFill(start_color="E0E0E0", fill_type="solid"))
    rule = Rule(type="expression", formula=["MOD(A1,2)=0"], dxf=dxf)
    ws.conditional_formatting.add("A1:C5", rule)

    wb.save(os.path.join(OUTPUT_DIR, "dxf_formula_rule.xlsx"))
    print("Created: dxf_formula_rule.xlsx")


def create_dxf_theme_color_test():
    """Test: DXF with theme colors (if supported by openpyxl version)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ThemeColor"

    for i in range(1, 6):
        ws.cell(row=i, column=1, value=i * 10)

    # Use theme color if available, fall back to RGB
    try:
        from openpyxl.styles.colors import Color as OpenpyxlColor

        font_color = OpenpyxlColor(theme=4)  # Accent 1
        dxf = DifferentialStyle(font=Font(bold=True, color=font_color))
    except (ImportError, TypeError):
        # Fall back to RGB if theme colors not supported
        dxf = DifferentialStyle(
            font=Font(bold=True, color="4472C4")  # Accent 1 approximation
        )

    rule = Rule(type="cellIs", operator="greaterThan", formula=["30"], dxf=dxf)
    ws.conditional_formatting.add("A1:A5", rule)

    wb.save(os.path.join(OUTPUT_DIR, "dxf_theme_color.xlsx"))
    print("Created: dxf_theme_color.xlsx")


def main():
    """Generate all DXF test fixtures."""
    print(f"Generating DXF test fixtures in: {OUTPUT_DIR}")
    print("-" * 50)

    # Basic tests
    create_dxf_font_basic_test()
    create_dxf_font_effects_test()
    create_dxf_fill_solid_test()
    create_dxf_border_test()

    # Alignment tests
    create_dxf_alignment_test()
    create_dxf_alignment_wrap_test()

    # Protection test
    create_dxf_protection_test()

    # Combined test
    create_dxf_full_style_test()

    # Advanced tests
    create_dxf_multiple_rules_test()
    create_dxf_formula_rule_test()
    create_dxf_theme_color_test()

    print("-" * 50)
    print("All fixtures generated successfully!")


if __name__ == "__main__":
    main()
